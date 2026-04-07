import io
import requests
import re  # <--- 새로 추가됨 (철근 텍스트 변환용)
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter  # <--- 새로 추가됨 (열 알파벳 계산용)
from datetime import datetime

app = Flask(__name__)
CORS(app)

TEMPLATE_URL = "https://hongstar2776-cmyk.github.io/My-Dashboard/resource/template_estmate.xlsx"
# <--- 새로 추가됨 (기초/지중보 전용 템플릿)
FOUNDATION_TEMPLATE_URL = "https://hongstar2776-cmyk.github.io/My-Dashboard/resource/template_foundation.xlsx" 

def write_row(ws, row_idx, data):
    ws[f'A{row_idx}'] = data.get('category') or ''
    ws[f'B{row_idx}'] = data.get('name') or ''
    ws[f'C{row_idx}'] = data.get('spec') or ''
    ws[f'D{row_idx}'] = data.get('unit') or ''

    if data.get('_type') == 'header':
        ws[f'A{row_idx}'].font = Font(bold=True)
        fill_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            cell = ws[f'{col}{row_idx}']
            cell.font = Font(bold=True)
            cell.fill = fill_blue
        return

    qty = float(data.get('qty') or 0)
    mat_up = float(data.get('mat_up') or 0)
    lab_up = float(data.get('lab_up') or 0)
    exp_up = float(data.get('exp_up') or 0)

    ws[f'E{row_idx}'] = qty
    ws[f'F{row_idx}'] = mat_up
    ws[f'H{row_idx}'] = lab_up
    ws[f'J{row_idx}'] = exp_up
    
    ws[f'G{row_idx}'] = f"=E{row_idx}*F{row_idx}"
    ws[f'I{row_idx}'] = f"=E{row_idx}*H{row_idx}"
    ws[f'K{row_idx}'] = f"=E{row_idx}*J{row_idx}"
    ws[f'L{row_idx}'] = f"=F{row_idx}+H{row_idx}+J{row_idx}"
    ws[f'M{row_idx}'] = f"=G{row_idx}+I{row_idx}+K{row_idx}"
    ws[f'N{row_idx}'] = data.get('note') or ''

def write_subtotal(ws, row_idx, cat_ranges, category):
    ws[f'A{row_idx}'] = category
    ws[f'B{row_idx}'] = f"[{category} 소계]"
    
    if cat_ranges:
        g_parts = [f"G{s}:G{e}" for s,e in cat_ranges]
        i_parts = [f"I{s}:I{e}" for s,e in cat_ranges]
        k_parts = [f"K{s}:K{e}" for s,e in cat_ranges]
        m_parts = [f"M{s}:M{e}" for s,e in cat_ranges]

        ws[f'G{row_idx}'] = f"=SUM({','.join(g_parts)})"
        ws[f'I{row_idx}'] = f"=SUM({','.join(i_parts)})"
        ws[f'K{row_idx}'] = f"=SUM({','.join(k_parts)})"
        ws[f'M{row_idx}'] = f"=SUM({','.join(m_parts)})"
    else:
        ws[f'G{row_idx}'] = "=0"
        ws[f'I{row_idx}'] = "=0"
        ws[f'K{row_idx}'] = "=0"
        ws[f'M{row_idx}'] = "=0"

    ws[f'A{row_idx}'].font = Font(bold=True)
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        cell = ws[f'{col}{row_idx}']
        cell.font = Font(bold=True)
        cell.fill = fill_gray

# =========================================================
# 🔹 기존 페이지에서 잘 쓰고 있는 API (건드리지 않음)
# =========================================================
@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        payload = request.json
        tabs = payload.get('tabs', [])
        meta = payload.get('meta', {}) 

        if not tabs:
            return jsonify({"error": "출력할 데이터가 없습니다."}), 400

        response = requests.get(TEMPLATE_URL)
        response.raise_for_status()
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        
        if "견적서겉표지" in wb.sheetnames:
            ws_cover = wb["견적서겉표지"]
            ws_cover['B10'] = meta.get('projectName', '')
            ws_cover['B13'] = meta.get('projectLocation', '')
            
        if "갑지" in wb.sheetnames:
            ws_gap = wb["갑지"]
            ws_gap['P1'] = meta.get('estimateDate', '')
            ws_gap['P2'] = meta.get('clientName', '')

        ws_total_summary = wb["총괄합계표"] if "총괄합계표" in wb.sheetnames else None
                
        if ws_total_summary and meta.get('documentTitle'):
            ws_total_summary['P1'] = meta.get('documentTitle').strip()
        
        base_est_sheet = wb["내역서"] if "내역서" in wb.sheetnames else wb.worksheets[0]
        base_sum_sheet = wb["공종별합계표"] if "공종별합계표" in wb.sheetnames else None
        
        for i, tab in enumerate(tabs):
            raw_tab_name = tab.get('name', f'내역서 {i+1}')
            
            clean_tab_name = raw_tab_name.replace("내역서 ", "").replace("내역서", "").strip()
            if not clean_tab_name: clean_tab_name = str(i+1)
            
            est_sheet_title = f"내역서({clean_tab_name})"
            sum_sheet_title = f"공종별합계표({clean_tab_name})"
            
            tab_data = tab.get('data', [])
            
            new_sum_sheet = None
            if base_sum_sheet:
                new_sum_sheet = wb.copy_worksheet(base_sum_sheet)
                new_sum_sheet.title = sum_sheet_title
                new_sum_sheet['A1'] = f" 공 종 별 합 계 표 ({clean_tab_name})"

            new_est_sheet = wb.copy_worksheet(base_est_sheet)
            new_est_sheet.title = est_sheet_title
            new_est_sheet.print_title_rows = '1:4'
            new_est_sheet['A1'] = f"내  역  서 ({clean_tab_name})"

            groups = {}
            for r in tab_data:
                cat = str(r.get('category') or '').strip() or '미지정'
                if cat not in groups:
                    groups[cat] = []
                groups[cat].append(r)

            current_row = 5 
            summary_data = [] 
            
            for cat, rows in groups.items():
                cat_ranges = [] 
                items_to_print = [{'_type': 'header', 'category': cat, 'name': cat}] + rows
                
                while items_to_print:
                    start_chunk_row = current_row
                    
                    if len(items_to_print) <= 20:
                        page_items = items_to_print[:]
                        items_to_print = []
                        
                        first_data = -1
                        last_data = -1
                        
                        for r in page_items:
                            write_row(new_est_sheet, current_row, r)
                            if r.get('_type') != 'header': 
                                if first_data == -1: first_data = current_row
                                last_data = current_row
                            current_row += 1
                            
                        if first_data != -1:
                            cat_ranges.append((first_data, last_data))
                        
                        while current_row < start_chunk_row + 20:
                            current_row += 1
                            
                        current_row += 1 
                        write_subtotal(new_est_sheet, current_row, cat_ranges, cat)
                        current_row += 2 
                        
                    elif len(items_to_print) == 21:
                        page_items = items_to_print[:21]
                        items_to_print = []
                        
                        first_data = -1
                        last_data = -1
                        
                        for r in page_items:
                            write_row(new_est_sheet, current_row, r)
                            if r.get('_type') != 'header':
                                if first_data == -1: first_data = current_row
                                last_data = current_row
                            current_row += 1
                        
                        if first_data != -1:
                            cat_ranges.append((first_data, last_data))
                        
                        write_subtotal(new_est_sheet, current_row, cat_ranges, cat)
                        current_row += 2 
                        
                    else: 
                        page_items = items_to_print[:21]
                        items_to_print = items_to_print[21:]
                        
                        first_data = -1
                        last_data = -1
                        
                        for r in page_items:
                            write_row(new_est_sheet, current_row, r)
                            if r.get('_type') != 'header':
                                if first_data == -1: first_data = current_row
                                last_data = current_row
                            current_row += 1
                            
                        if first_data != -1:
                            cat_ranges.append((first_data, last_data))
                        
                        new_est_sheet[f'B{current_row}'] = "[...다음 장으로 이어짐]"
                        current_row += 1
                        new_est_sheet[f'B{current_row}'] = "[...다음 장으로 이어짐]"
                        current_row += 1
                
                summary_data.append({
                    'category': cat,
                    'ranges': cat_ranges
                })
            
            last_row = current_row - 1
            new_est_sheet.print_area = f"B1:N{last_row}"
            
            total_row_for_summary = 0 
            
            if new_sum_sheet:
                sum_row = 5
                for s_data in summary_data:
                    cat = s_data['category']
                    ranges = s_data['ranges']
                    
                    new_sum_sheet[f'A{sum_row}'] = cat
                    new_sum_sheet[f'B{sum_row}'] = cat
                    
                    if ranges:
                        est_sheet_name = new_est_sheet.title
                        g_parts = [f"'{est_sheet_name}'!G{s}:G{e}" for s,e in ranges]
                        i_parts = [f"'{est_sheet_name}'!I{s}:I{e}" for s,e in ranges]
                        k_parts = [f"'{est_sheet_name}'!K{s}:K{e}" for s,e in ranges]
                        m_parts = [f"'{est_sheet_name}'!M{s}:M{e}" for s,e in ranges]
                        
                        new_sum_sheet[f'G{sum_row}'] = f"=SUM({','.join(g_parts)})"
                        new_sum_sheet[f'I{sum_row}'] = f"=SUM({','.join(i_parts)})"
                        new_sum_sheet[f'K{sum_row}'] = f"=SUM({','.join(k_parts)})"
                        new_sum_sheet[f'M{sum_row}'] = f"=SUM({','.join(m_parts)})"
                    
                    for col in ['A', 'B', 'G', 'I', 'K', 'M']:
                        new_sum_sheet[f'{col}{sum_row}'].font = Font(bold=True)
                        
                    sum_row += 1

                num_categories = len(summary_data)
                calc_last_data_row = max(24, 4 + num_categories)
                
                if num_categories > 20:
                    for r_idx in range(25, 28):
                        for c_idx in ['A', 'B', 'G', 'I', 'K', 'M', 'N']:
                            new_sum_sheet[f'{c_idx}{r_idx}'] = None

                total_row_for_summary = calc_last_data_row + 2
                print_end_row = calc_last_data_row + 3

                new_sum_sheet[f'B{total_row_for_summary}'] = "[합 계]"
                new_sum_sheet[f'B{total_row_for_summary}'].font = Font(bold=True)

                new_sum_sheet[f'G{total_row_for_summary}'] = f"=SUM(G5:G{calc_last_data_row})"
                new_sum_sheet[f'I{total_row_for_summary}'] = f"=SUM(I5:I{calc_last_data_row})"
                new_sum_sheet[f'K{total_row_for_summary}'] = f"=SUM(K5:K{calc_last_data_row})"
                new_sum_sheet[f'M{total_row_for_summary}'] = f"=SUM(M5:M{calc_last_data_row})"

                for col in ['G', 'I', 'K', 'M']:
                    new_sum_sheet[f'{col}{total_row_for_summary}'].font = Font(bold=True)

                new_sum_sheet.print_area = f"B1:N{print_end_row}"

            if ws_total_summary and new_sum_sheet and total_row_for_summary > 0:
                t_row = 5 + i 
                ws_total_summary[f'B{t_row}'] = raw_tab_name
                ws_total_summary[f'G{t_row}'] = f"='{sum_sheet_title}'!G{total_row_for_summary}" 
                ws_total_summary[f'I{t_row}'] = f"='{sum_sheet_title}'!I{total_row_for_summary}" 
                ws_total_summary[f'K{t_row}'] = f"='{sum_sheet_title}'!K{total_row_for_summary}" 
                ws_total_summary[f'M{t_row}'] = f"='{sum_sheet_title}'!M{total_row_for_summary}" 

        if base_est_sheet:
            base_est_sheet.sheet_state = 'hidden'
        if base_sum_sheet:
            base_sum_sheet.sheet_state = 'hidden'

        for idx, sheet_name in enumerate(wb.sheetnames):
            if wb[sheet_name].sheet_state == 'visible':
                wb.active = idx
                break

        project_name = meta.get('projectName', '000신축공사')
        estimate_date = meta.get('estimateDate', '000년 00월')
        today_date = datetime.now().strftime("%Y%m%d") 
        
        final_filename = f"견적서_{project_name}_{estimate_date}_{today_date}.xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, 
            as_attachment=True, 
            download_name=final_filename, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("서버 에러 발생:", str(e))
        return jsonify({"error": str(e)}), 500


# =========================================================
# 🌟 새롭게 추가된 기초/지중보 산출서 API
# =========================================================
@app.route('/api/export_foundation', methods=['GET', 'POST'])
def export_foundation_excel():
    # GET 방식(주소창 접속) 생존 테스트용
    if request.method == 'GET':
        return "Foundation API 정상 작동 중!", 200

    try:
        payload = request.json
        project_name = payload.get('projectName', '000신축공사')
        items = payload.get('items', [])
        summary = payload.get('summary', {})

        if not items and not summary:
            return jsonify({"error": "출력할 데이터가 없습니다."}), 400

        # 템플릿 파일 다운로드 및 openpyxl 워크북 로드
        response = requests.get(FOUNDATION_TEMPLATE_URL)
        response.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        # [작업 1] "내역서" 시트 작성
        if "내역서" in wb.sheetnames:
            ws_summary = wb["내역서"]
            ws_summary['B2'] = project_name
            row_idx = 5
            
            for spec, qty in summary.get('concrete', {}).items():
                ws_summary[f'B{row_idx}'] = "레미콘"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "m3"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1
                
            for spec, qty in summary.get('formwork', {}).items():
                ws_summary[f'B{row_idx}'] = "거푸집"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "m2"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1
                
            for spec, qty in summary.get('rebar', {}).items():
                ws_summary[f'B{row_idx}'] = "철근"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "kg"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1

        # [작업 2] "상세산출서" 시트 작성
        if "상세산출서" in wb.sheetnames:
            ws_detail = wb["상세산출서"]
            start_row = 2
            
            for item in items:
                ws_detail[f'B{start_row}'] = item.get('type', '')
                ws_detail[f'C{start_row}'] = item.get('name', '')
                
                ws_detail[f'A{start_row+1}'] = "콘크리트"
                ws_detail[f'B{start_row+1}'] = item.get('conc', 0)
                ws_detail[f'C{start_row+1}'] = f"{item.get('fck', '')} MPa"
                ws_detail[f'D{start_row+1}'] = item.get('formulas', {}).get('conc', '')
                
                ws_detail[f'A{start_row+2}'] = "거푸집"
                ws_detail[f'B{start_row+2}'] = item.get('form', 0)
                ws_detail[f'C{start_row+2}'] = item.get('formulas', {}).get('form', '')
                
                ws_detail[f'A{start_row+3}'] = "철근"
                ws_detail[f'B{start_row+3}'] = item.get('rebarTotal', 0)
                ws_detail[f'C{start_row+3}'] = item.get('formulas', {}).get('rebar', '')
                ws_detail[f'D{start_row+3}'] = item.get('formulas', {}).get('details', '')
                
                rebar_map = item.get('rebarDetailsMap', {})
                col_idx = 5
                
                for k, v in sorted(rebar_map.items()):
                    col_letter = get_column_letter(col_idx)
                    match = re.match(r'(HD\d+)\((SD\d+)\)', k)
                    formatted_key = f"{match.group(2)},{match.group(1)}" if match else k
                    
                    ws_detail[f'{col_letter}{start_row+3}'] = f"{formatted_key} : {round(v, 3)} kg"
                    col_idx += 1
                
                start_row += 5

        # [작업 3] 엑셀 파일 저장 및 클라이언트 반환
        today_date = datetime.now().strftime("%Y%m%d") 
        final_filename = f"물량산출서_{project_name}_{today_date}.xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, 
            as_attachment=True, 
            download_name=final_filename, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("서버 에러 발생:", str(e))
        return jsonify({"error": str(e)}), 500
