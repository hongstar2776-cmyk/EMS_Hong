import io
import requests
import re
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
CORS(app)

TEMPLATE_URL = "https://hongstar2776-cmyk.github.io/My-Dashboard/resource/template_foundation.xlsx"

# [수정됨] GET 방식을 추가하여 브라우저 주소창에서도 생존 여부를 확인할 수 있게 만듦
@app.route('/api/export_foundation', methods=['GET', 'POST'])
def export_foundation_excel():
    # 💡 브라우저 주소창에 직접 쳤을 때 (GET 요청) 잘 돌아가고 있는지 확인하는 응답
    if request.method == 'GET':
        return "API 서버가 정상적으로 살아있습니다! (Foundation API)", 200

    try:
        # 프론트엔드(HTML)에서 보낸 JSON 데이터 파싱
        payload = request.json
        project_name = payload.get('projectName', '000신축공사')
        items = payload.get('items', [])
        summary = payload.get('summary', {})

        if not items and not summary:
            return jsonify({"error": "출력할 데이터가 없습니다."}), 400

        # 템플릿 파일 다운로드 및 openpyxl 워크북 로드
        response = requests.get(TEMPLATE_URL)
        response.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        # ==========================================
        # [작업 1] "내역서" 시트 (규격별 총괄집계) 작성
        # ==========================================
        if "내역서" in wb.sheetnames:
            ws_summary = wb["내역서"]
            
            # B2셀에 공사명 입력
            ws_summary['B2'] = project_name
            
            row_idx = 5
            
            # 레미콘 내역 쓰기
            for spec, qty in summary.get('concrete', {}).items():
                ws_summary[f'B{row_idx}'] = "레미콘"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "m3"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1
                
            # 거푸집 내역 쓰기
            for spec, qty in summary.get('formwork', {}).items():
                ws_summary[f'B{row_idx}'] = "거푸집"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "m2"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1
                
            # 철근 내역 쓰기
            for spec, qty in summary.get('rebar', {}).items():
                ws_summary[f'B{row_idx}'] = "철근"
                ws_summary[f'C{row_idx}'] = spec
                ws_summary[f'D{row_idx}'] = "kg"
                ws_summary[f'E{row_idx}'] = qty
                row_idx += 1

        # ==========================================
        # [작업 2] "상세산출서" 시트 (부재별 상세내역) 작성
        # ==========================================
        if "상세산출서" in wb.sheetnames:
            ws_detail = wb["상세산출서"]
            
            start_row = 2
            
            for item in items:
                # 1행: 구분 및 부재명
                ws_detail[f'B{start_row}'] = item.get('type', '')
                ws_detail[f'C{start_row}'] = item.get('name', '')
                
                # 2행: 레미콘 정보
                ws_detail[f'A{start_row+1}'] = "콘크리트"
                ws_detail[f'B{start_row+1}'] = item.get('conc', 0)
                ws_detail[f'C{start_row+1}'] = f"{item.get('fck', '')} MPa"
                ws_detail[f'D{start_row+1}'] = item.get('formulas', {}).get('conc', '')
                
                # 3행: 거푸집 정보
                ws_detail[f'A{start_row+2}'] = "거푸집"
                ws_detail[f'B{start_row+2}'] = item.get('form', 0)
                ws_detail[f'C{start_row+2}'] = item.get('formulas', {}).get('form', '')
                
                # 4행: 철근 정보 (총합, 산출식, 상세정보)
                ws_detail[f'A{start_row+3}'] = "철근"
                ws_detail[f'B{start_row+3}'] = item.get('rebarTotal', 0)
                ws_detail[f'C{start_row+3}'] = item.get('formulas', {}).get('rebar', '')
                ws_detail[f'D{start_row+3}'] = item.get('formulas', {}).get('details', '')
                
                # E열부터 우측으로 늘리면서 철근 규격별 수량 기입 (E, F, G, H...)
                rebar_map = item.get('rebarDetailsMap', {})
                col_idx = 5  # 5는 E열을 의미함
                
                for k, v in sorted(rebar_map.items()):
                    col_letter = get_column_letter(col_idx)
                    
                    # HD10(SD400) 형태를 SD400,HD10 형태로 포맷팅
                    match = re.match(r'(HD\d+)\((SD\d+)\)', k)
                    formatted_key = f"{match.group(2)},{match.group(1)}" if match else k
                    
                    # 엑셀 셀에 [규격 : 중량kg] 형태로 입력
                    ws_detail[f'{col_letter}{start_row+3}'] = f"{formatted_key} : {round(v, 3)} kg"
                    col_idx += 1
                
                # 다음 부재 작성을 위해 행을 5칸 건너뜀 (부재 간 1줄 띄우기)
                start_row += 5

        # ==========================================
        # [작업 3] 엑셀 파일 저장 및 클라이언트 반환
        # ==========================================
        today_date = datetime.now().strftime("%Y%m%d") 
        final_filename = f"물량산출서_{project_name}_{today_date}.xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, 
            as_attachment=True, 
            download_name=final_filename, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("서버 에러 발생:", str(e))
        return jsonify({"error": str(e)}), 500
